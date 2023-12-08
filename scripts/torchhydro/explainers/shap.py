"""
Author: Wenyu Ouyang
Date: 2023-10-19 21:34:29
LastEditTime: 2023-10-28 13:00:52
LastEditors: Wenyu Ouyang
Description: SHAP methods for deep learning models
FilePath: \torchhydro\torchhydro\explainers\shap.py
Copyright (c) 2023-2024 Wenyu Ouyang. All rights reserved.
"""
from typing import List
import torch
import numpy as np
import shap
import matplotlib.pyplot as plt
import seaborn as sns
import os
import pandas as pd
from openpyxl import Workbook, load_workbook


def plot_summary_shap_values(shap_values: torch.tensor, columns):
    mean_shap_values = shap_values.mean(dim=(0, 1))

    # Tensor to NumPy array to use matplotlib and seaborn
    mean_shap_values = mean_shap_values.abs().mean(dim=0).numpy()

    # seaborn barplot
    plt.figure(figsize=(10, 8))
    sns.barplot(x=mean_shap_values, y=columns, order=columns[::-1], orient="h")
    plt.xlabel("Mean Absolute SHAP Value")
    plt.title("Summary Plot of SHAP Values")
    plt.show()


def plot_summary_shap_values_over_time_series(shap_values: torch.tensor, columns):
    # Compute the absolute mean SHAP values over batches
    abs_mean_shap_values = shap_values.mean(dim=1).abs()

    # Further compute the mean over observations
    multi_shap_values = abs_mean_shap_values.mean(dim=1).numpy()

    # Plotting
    plt.figure(figsize=(10, 8))

    bottom = np.zeros(len(columns))
    for i, pred_shap_values in enumerate(multi_shap_values):
        plt.barh(columns[::-1], pred_shap_values, label=f"time-step {i}", left=bottom)
        bottom += pred_shap_values

    plt.xlabel("Mean Absolute SHAP Value Over Time Steps")
    plt.title("Summary Plot of SHAP Values Over Time Series")
    plt.legend()
    plt.show()


def jitter(values, jitter_strength=0.005):
    """Add some jitter to the values."""
    return values + jitter_strength * np.random.randn(*values.shape)


def plot_shap_values_from_history(shap_values: torch.tensor, history: torch.tensor):
    # Compute mean values across the specified axes
    mean_shap_values = shap_values.mean(dim=[0, 1])
    mean_history_values = history.mean(dim=1)
    figs: List[plt.Figure] = []
    # Loop through features and plot each scatter plot
    for feature_history, feature_shap_values in zip(
        mean_history_values.unbind(0), mean_shap_values.unbind(0)
    ):
        fig, ax = plt.subplots(figsize=(8, 6))
        # Convert tensor to numpy array for plotting
        feature_shap_values_np = feature_shap_values.numpy()
        feature_history_np = feature_history.numpy()
        scatter = ax.scatter(
            feature_shap_values_np,
            jitter(feature_shap_values_np),
            c=feature_history_np,
            cmap="bwr",  # Blue to Red color map
            marker="o",
        )
        ax.set_ylim(-0.05, 0.05)
        ax.set_xlabel("shap value")
        ax.set_yticks([])  # Hide y-axis ticks as they represent jittered values
        # Colorbar configuration
        cbar = plt.colorbar(scatter, ax=ax)
        cbar.set_label("feature values")
        figs.append(fig)
    return figs


def deep_explain_model_summary_plot(dl_model, test_dataset) -> None:
    """Generate feature summary plot for trained deep learning models

    Parameters
    ----------
    model
        trained model
    test_dataset
        test dataset
    """
    dl_model.eval()
    history = test_dataset.__getitem__(0)[0].unsqueeze(0)
    dl_model = dl_model.to("cpu")
    deep_explainer = shap.DeepExplainer(dl_model, history)
    shap_values = shap_results(deep_explainer, history)
    # summary plot shows overall feature ranking
    # by average absolute shap values
    fig = plot_summary_shap_values(shap_values, test_dataset.df.columns)
    abs_mean_shap_values = shap_values.mean(axis=["preds", "batches"])
    multi_shap_values = abs_mean_shap_values.mean(axis="observations")
    # summary plot for multi-step outputs
    # multi_shap_values = shap_values.apply_along_axis(np.mean, 'batches')
    fig = plot_summary_shap_values_over_time_series(
        shap_values, test_dataset.df.columns
    )
    history_numpy = torch.tensor(
        history.cpu().numpy(), names=["batches", "observations", "features"]
    )

    shap_values = shap_results(deep_explainer, history)
    figs = plot_shap_values_from_history(shap_values, history_numpy)


def shap_results(deep_explainer, history):
    result = deep_explainer.shap_values(history)
    result = np.stack(result)
    # shap_values needs to be 4-dimensional
    if len(result.shape) != 4:
        result = np.expand_dims(result, axis=0)
    result = torch.tensor(
        result, names=["preds", "batches", "observations", "features"]
    )

    return result


def plot_shap_value_heatmaps(shap_values: torch.tensor):
    # Compute the average shap values over batches
    average_shap_value_over_batches = shap_values.mean(dim=shap_values.dim() - 1)

    figs: List[plt.Figure] = []

    # Extracting the shapes of the data for the x and y ticks
    x = list(range(average_shap_value_over_batches.size(-1)))
    y = list(range(average_shap_value_over_batches.size(0)))

    # Iterate through features and plot each heatmap
    for shap_values_features in average_shap_value_over_batches.unbind(0):
        fig, ax = plt.subplots(figsize=(8, 6))

        # Convert tensor to numpy array for plotting
        shap_values_features_np = shap_values_features.numpy()

        # Plotting heatmap
        cax = ax.matshow(shap_values_features_np, cmap="bwr", aspect="auto")

        ax.set_xticks(x)
        ax.set_yticks(y)
        ax.set_xlabel("sequence history steps")
        ax.set_ylabel("prediction steps")

        # Adding colorbar to the right of the heatmap
        cbar = fig.colorbar(cax, ax=ax)
        cbar.set_label("feature values")

        figs.append(fig)

    return figs


def deep_explain_model_heatmap(dl_model, test_dataset) -> None:
    """Generate feature heatmap for prediction at a start time
    Parameters
    ----------
    model
        trained model
    test_dataset
        test dataset
    Returns
    -------
    None
    """
    dl_model.eval()
    history = test_dataset.__getitem__(0)[0]
    # background shape (L, N, M)
    # L - batch size, N - history length, M - feature size
    # for each element in each N x M batch in L,
    # attribute to each prediction in forecast len
    deep_explainer = shap.DeepExplainer(dl_model, history)
    shap_values = shap_results(deep_explainer, history)
    figs = plot_shap_value_heatmaps(shap_values)


def shap_summary_plot(dl_model, train_dataset, test_dataset) -> None:
    dl_model.eval()

    history_list = []
    for i in range(
        min(
            168,
            int(train_dataset.num_samples / len(train_dataset.data_cfgs["object_ids"])),
        )
    ):
        h = train_dataset.__getitem__(i)[0]
        history_list.append(h)
    history = torch.cat(history_list).cuda()

    test_list = []
    for i in range(
        min(
            1,
            int(test_dataset.num_samples / len(test_dataset.data_cfgs["object_ids"])),
        )
    ):
        t = test_dataset.__getitem__(i)[0]
        test_list.append(t)
    test = torch.cat(test_list).cuda()
    print(test.shape)
    dl_model = dl_model.cuda()
    torch.backends.cudnn.enabled = False
    e = shap.DeepExplainer(dl_model, history)
    shap_values = e.shap_values(test)[0]

    shap_values_avg = (
        shap_values.squeeze(2)
        .reshape(shap_values.shape[0], shap_values.shape[1], -1)
        .mean(axis=-2)
    )
    test_tensor_avg = (
        test.squeeze(2).reshape(test.shape[0], test.shape[1], -1).mean(axis=-2)
    ).to("cpu")
    """
    np.save(
        os.path.join(test_dataset.data_cfgs["test_path"], "shap_values_avg.npy"),
        shap_values_avg,
    )
    """
    if not os.path.exists(test_dataset.data_cfgs["shap_path"]):
        os.makedirs(test_dataset.data_cfgs["shap_path"])
    """
    np.savetxt(
        os.path.join(test_dataset.data_cfgs["shap_path"], "shap_values_avg.xlsx"),
        shap_values_avg,
        delimiter=",",
    )
    """
    df = pd.DataFrame(shap_values_avg)
    xlsx_path = test_dataset.data_cfgs["shap_path"]
    if not os.path.exists(xlsx_path):
        os.makedirs(xlsx_path)

    lonlat_file = test_dataset.data_cfgs["model_path"] + "/lonlat.txt"

    with open(lonlat_file, "r") as file:
        lines = file.readlines()

    # 假设每一行的元素都是通过空格分隔的
    line1 = lines[0].strip().split()
    line2 = lines[1].strip().split()

    # 生成所有可能的组合
    combinations = []
    for item1 in line1:
        for item2 in line2:
            combinations.append((item1, item2))

    df.to_excel(
        os.path.join(
            test_dataset.data_cfgs["shap_path"]
            + "/"
            + test_dataset.data_cfgs["shap_name"]
            + ".xlsx"
        ),
        index=False,
    )
    xlsx_file_name = os.path.join(
        test_dataset.data_cfgs["shap_path"]
        + "/"
        + test_dataset.data_cfgs["shap_name"]
        + ".xlsx"
    )
    if not os.path.exists(xlsx_file_name):
        wb = Workbook()
        wb.save(xlsx_file_name)

    # 加载工作簿
    wb = load_workbook(xlsx_file_name)
    sheet = wb.active

    # 将列表中的元素填充到第一行的不同列
    for col, value in enumerate(combinations, start=1):
        sheet.cell(row=1, column=col, value=str(value))

    # 保存并关闭工作簿
    wb.save(xlsx_file_name)
    wb.close()

    torch.save(
        test_tensor_avg,
        os.path.join(test_dataset.data_cfgs["test_path"], "test_tensor_avg.pth"),
    )
    # shap.summary_plot(shap_values_avg, test_tensor_avg)
    shap.summary_plot(shap_values_avg, test_tensor_avg, show=False)
    plt.savefig(
        os.path.join(
            test_dataset.data_cfgs["shap_path"]
            + "/"
            + test_dataset.data_cfgs["shap_name"]
            + ".png"
        ),
        transparent=True,
    )
